import urllib.request
import re
import openpyxl as xl
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference


class AppURLopener(urllib.request.FancyURLopener):
    version = "Mozilla/5.0"


def setting_up_source_code(link):
    url = link
    opener = AppURLopener()
    page = opener.open(url)
    html = page.read().decode()
    return html

# the links in the source code are short links, so this function adds gitHub.com to all the links,
# so you can open them just from clicking on it


def add_git_to_links(list_of_links):
    counter = 0
    https = 'https://github.com'
    for link in list_of_links:
        link = https + link
        list_of_links[counter] = link
        counter = counter + 1
    return list_of_links


def populate_excel(list, num_column, file_name, is_number=False):
    if is_number:
        populate_excel_numbers(list, num_column, file_name)
        return
    counter = 0
    for value in range(2, len(list)+1):
        ws.cell(value, num_column).value = list[counter]
        counter += 1
    wb.save(file_name)


def populate_excel_numbers(list, num_column, file_name):
    counter = 0
    for value in range(2, len(list)+1):
        if (list[counter])[-1] == 'k':
            # in this line I am getting the number without the letter k, converting it to a float
            # and multiplying by 1000, which is what k means
            number = float((list[counter])[0:-1]) * 1000
            list[counter] = number
        ws.cell(value, num_column).value = int(list[counter])
        counter += 1
    wb.save(file_name)


def find_values(filepath):
    info = re.compile(filepath)
    result = re.findall(info, html)
    return result


def setting_up_column_headings(list, filename):
    # list_headings = ['Titles', 'Last Updated', 'Language', 'Stars', 'Links']
    counter = 0
    for value in range(1, len(list)+1):
        # first value is row, second column
        ws.cell(1, value).value = list[counter]
        ws.cell(1, value).font = Font(bold=True)
        counter += 1
    wb.save(filename)


html = setting_up_source_code('https://github.com/search?q=game')
game_names = find_values('''<p class="mb-1">
        (.*)
      </p>''')
# print(titles)
# print(len(titles))
# string = re.compile('''<p class="mb-1">
#         (.*)
#       </p>''')
# result = re.findall(string, html)
# print(result)
lastUpdated = find_values('''<div class="mr-3">
            Updated <.*>(.*)</relative-time>
          </div>''')
# print(lastUpdated)
# print(len(lastUpdated))
# Update = re.compile('''<div class="mr-3">
#             Updated <.*>(.*)</relative-time>
#           </div>''')
# resultUpdate = re.findall(Update, html)
# print(resultUpdate)
language = find_values(''' <span itemprop="programmingLanguage">(.*)</span>''')
# print(language)
# print(len(language))
# Language = re.compile(''' <span itemprop="programmingLanguage">(.*)</span>''')
# resultLanguage = re.findall(Language, html)
# print(resultLanguage)
stars = find_values(''' <path d=.*></path>
</svg>
              (.*)
            </a>
          </div>''')
# print(stars)
# print(len(stars))
# Starts = re.compile(''' <path d=.*></path>
# </svg>
#               (.*)
#             </a>
#           </div>''')
# resultStarts = re.findall(Starts, html)
# print(resultStarts)
links = add_git_to_links(find_values('''<a class="Link--muted" href="(.*)/stargazers">'''))
# print(links)
# print(len(links))
# Link = re.compile('''<a class="Link--muted" href="(.*)/stargazers">''')
# resultLink = re.findall(Link, html)
# counter = 0
# for link in resultLink:
#     https = 'https://github.com'
#     link = https + link
#     resultLink[counter] = link
#     counter = counter + 1
# print(resultLink)
wb = xl.load_workbook('ShyndeeMandelPythonWebScraping.xlsx')
ws = wb.active
# list_headings = ['Titles', 'Last Updated',  'Language', 'Stars', 'Links']
# counter = 0
# for row_headings in range(1, len(list_headings)):
#     # first value is row, second column
#     ws.cell(1, row_headings).value = list_headings[counter]
#     counter += 1
# wb.save('ShyndeeMandelPythonWebScraping.xlsx')
list_headings = ['Titles', 'Last Updated', 'Language', 'Stars', 'Links']
setting_up_column_headings(list_headings, 'ShyndeeMandelPythonWebScraping.xlsx')
populate_excel(game_names, 1, 'ShyndeeMandelPythonWebScraping.xlsx')
# counter = 0
# for value in range(2, len(titles)):
#     ws.cell(value, 1).value = titles[counter]
#     counter += 1
# wb.save('ShyndeeMandelPythonWebScraping.xlsx')

# counter = 0
# for value in range(2, len(lastUpdated)):
#     ws.cell(value, 2).value = lastUpdated[counter]
#     counter += 1
# wb.save('ShyndeeMandelPythonWebScraping.xlsx')
populate_excel(lastUpdated, 2, 'ShyndeeMandelPythonWebScraping.xlsx')
# counter = 0
# for value in range(2, len(language)):
#     ws.cell(value, 3).value = language[counter]
#     counter += 1
# wb.save('ShyndeeMandelPythonWebScraping.xlsx')
populate_excel(language, 3, 'ShyndeeMandelPythonWebScraping.xlsx')
# counter = 0
# for value in range(2, len(stars)):
#     ws.cell(value, 4).value = stars[counter]
#     counter += 1
# wb.save('ShyndeeMandelPythonWebScraping.xlsx')
populate_excel(stars, 4, 'ShyndeeMandelPythonWebScraping.xlsx', True)
populate_excel(links, 5, 'ShyndeeMandelPythonWebScraping.xlsx')

# create the chart
bar_chart = BarChart()
bar_chart.type = 'bar'
bar_chart.style = 7
bar_chart.title = 'Game Popularity'
bar_chart.width = 45.15
bar_chart.height = 16.68


data = Reference(worksheet=ws, min_row=1, max_row=10, min_col=4, max_col=4)
cats = Reference(ws, min_col=1, min_row=2, max_row=10)
bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(cats)
bar_chart.shape = 4
ws.add_chart(bar_chart, 'A14')

wb.save('ShyndeeMandelPythonWebScraping.xlsx')
